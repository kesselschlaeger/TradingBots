"""Tearsheet-Reporter: Sharpe, MaxDD, CAGR, Win-Rate, Expectancy.

Erweiterungen:
  - build_exit_reason_stats / format_exit_reason_stats: Exit-Grund-Analyse
  - export_trades: CSV + Excel-Export (openpyxl)
  - format_enriched_tearsheet: MAE/MFE-Summary, Top/Worst-Trades
"""
from __future__ import annotations

from dataclasses import asdict, dataclass, fields as dc_fields
from datetime import datetime
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd

from core.logging import get_logger
from core.models import EnrichedTrade, Trade

log = get_logger(__name__)


@dataclass
class Tearsheet:
    initial_capital: float
    final_equity: float
    total_return_pct: float
    cagr_pct: float
    max_drawdown_pct: float
    sharpe: float
    sortino: float
    win_rate: float
    avg_win: float
    avg_loss: float
    profit_factor: float
    expectancy: float
    num_trades: int

    def as_dict(self) -> dict:
        return asdict(self)


def _annualization_factor(equity: pd.Series) -> float:
    if len(equity) < 2:
        return 252.0
    span_days = (equity.index[-1] - equity.index[0]).days
    if span_days <= 0:
        return 252.0
    bars_per_day = len(equity) / max(span_days, 1)
    return bars_per_day * 252.0


def _max_drawdown(equity: pd.Series) -> float:
    if equity.empty:
        return 0.0
    rolling_max = equity.cummax()
    dd = (equity - rolling_max) / rolling_max.replace(0, np.nan)
    return float(abs(dd.min())) if not dd.isna().all() else 0.0


def _sharpe(returns: pd.Series, ann: float, rf: float = 0.0) -> float:
    if returns.std(ddof=0) == 0 or len(returns) < 2:
        return 0.0
    excess = returns - rf / ann
    return float(np.sqrt(ann) * excess.mean() / returns.std(ddof=0))


def _sortino(returns: pd.Series, ann: float, rf: float = 0.0) -> float:
    downside = returns[returns < 0]
    if len(downside) < 2 or downside.std(ddof=0) == 0:
        return 0.0
    excess = returns - rf / ann
    return float(np.sqrt(ann) * excess.mean() / downside.std(ddof=0))


def _trade_pnls(trades: list[Trade]) -> list[float]:
    return [float(t.pnl) for t in trades if t.pnl != 0.0]


def build_tearsheet(equity: pd.Series, trades: list[Trade],
                    initial_capital: float) -> Tearsheet:
    if equity.empty:
        return Tearsheet(initial_capital, initial_capital, 0, 0, 0, 0, 0,
                         0, 0, 0, 0, 0, 0)

    final = float(equity.iloc[-1])
    total_ret = (final / initial_capital - 1.0) * 100.0

    span_days = max((equity.index[-1] - equity.index[0]).days, 1)
    cagr = ((final / initial_capital) ** (365.0 / span_days) - 1.0) * 100.0 \
        if final > 0 else -100.0

    returns = equity.pct_change().dropna()
    ann = _annualization_factor(equity)

    pnls = _trade_pnls(trades)
    wins = [p for p in pnls if p > 0]
    losses = [p for p in pnls if p < 0]
    win_rate = (len(wins) / len(pnls)) if pnls else 0.0
    avg_win = float(np.mean(wins)) if wins else 0.0
    avg_loss = float(np.mean(losses)) if losses else 0.0
    pf = (sum(wins) / abs(sum(losses))) if losses and sum(losses) < 0 else 0.0
    expectancy = (win_rate * avg_win + (1 - win_rate) * avg_loss) if pnls else 0.0

    return Tearsheet(
        initial_capital=float(initial_capital),
        final_equity=final,
        total_return_pct=total_ret,
        cagr_pct=cagr,
        max_drawdown_pct=_max_drawdown(equity) * 100.0,
        sharpe=_sharpe(returns, ann),
        sortino=_sortino(returns, ann),
        win_rate=win_rate,
        avg_win=avg_win,
        avg_loss=avg_loss,
        profit_factor=pf,
        expectancy=expectancy,
        num_trades=len(pnls),
    )


def _fmt_ts(ts: Optional[datetime]) -> str:
    if ts is None:
        return "—"
    return ts.strftime("%Y-%m-%d %H:%M %Z").strip() or ts.strftime("%Y-%m-%d %H:%M")


def format_tearsheet(
    ts: Tearsheet,
    *,
    start_ts: Optional[datetime] = None,
    end_ts: Optional[datetime] = None,
    strategy_name: str = "",
    allow_shorts: Optional[bool] = None,
    mit_enabled: Optional[bool] = None,
    enriched_trades: Optional[list[EnrichedTrade]] = None,
) -> str:
    header = ""
    if strategy_name:
        header += f"Strategy:       {strategy_name}\n"
    if start_ts is not None or end_ts is not None:
        span_days = ""
        if start_ts is not None and end_ts is not None:
            days = max((end_ts - start_ts).days, 0)
            span_days = f"  ({days} d)"
        header += f"Period:         {_fmt_ts(start_ts)} → {_fmt_ts(end_ts)}{span_days}\n"
    if allow_shorts is not None:
        header += f"Shorts allowed: {'yes' if allow_shorts else 'no'}\n"
    if mit_enabled is not None:
        header += f"MIT overlay:    {'on' if mit_enabled else 'off'}\n"
    if header:
        header += "-" * 50 + "\n"
    base = header + (
        f"Initial:        ${ts.initial_capital:,.2f}\n"
        f"Final:          ${ts.final_equity:,.2f}\n"
        f"Total Return:   {ts.total_return_pct:+.2f}%\n"
        f"CAGR:           {ts.cagr_pct:+.2f}%\n"
        f"Max Drawdown:   {ts.max_drawdown_pct:.2f}%\n"
        f"Sharpe:         {ts.sharpe:.2f}\n"
        f"Sortino:        {ts.sortino:.2f}\n"
        f"Trades:         {ts.num_trades}\n"
        f"Win Rate:       {ts.win_rate:.1%}\n"
        f"Avg Win:        ${ts.avg_win:,.2f}\n"
        f"Avg Loss:       ${ts.avg_loss:,.2f}\n"
        f"Profit Factor:  {ts.profit_factor:.2f}\n"
        f"Expectancy:     ${ts.expectancy:,.2f}/Trade\n"
    )
    if enriched_trades:
        base += _format_mae_mfe_summary(enriched_trades)
        base += _format_top_worst_trades(enriched_trades)
    return base


# ═══════════════════════════════════════════════════════════════════════════
#  Exit-Reason-Statistik
# ═══════════════════════════════════════════════════════════════════════════

def build_exit_reason_stats(trades: list[EnrichedTrade]) -> pd.DataFrame:
    """Aggregiert abgeschlossene Trades nach Exit-Grund.

    Sortiert nach count descending.
    """
    if not trades:
        return pd.DataFrame()

    rows = []
    for t in trades:
        rows.append({
            "exit_reason": t.exit_reason,
            "pnl_net": t.pnl_net,
            "r_multiple": t.r_multiple,
            "hold_days": t.hold_days,
            "pnl_pct": t.pnl_pct,
        })
    df = pd.DataFrame(rows)
    total = len(df)

    stats = []
    for reason, grp in df.groupby("exit_reason", sort=False):
        cnt = len(grp)
        win_count = int((grp["pnl_net"] > 0).sum())
        stats.append({
            "exit_reason": reason,
            "count": cnt,
            "count_pct": cnt / total * 100.0 if total else 0.0,
            "win_count": win_count,
            "win_rate_pct": win_count / cnt * 100.0 if cnt else 0.0,
            "pnl_total": grp["pnl_net"].sum(),
            "pnl_mean": grp["pnl_net"].mean(),
            "pnl_median": grp["pnl_net"].median(),
            "r_mean": grp["r_multiple"].mean(),
            "r_median": grp["r_multiple"].median(),
            "hold_days_mean": grp["hold_days"].mean(),
            "pnl_pct_mean": grp["pnl_pct"].mean(),
        })

    result = pd.DataFrame(stats)
    return result.sort_values("count", ascending=False).reset_index(drop=True)


def _build_interpretations(stats_df: pd.DataFrame) -> list[str]:
    """Heuristische Interpretation der Exit-Statistik."""
    hints: list[str] = []
    for _, row in stats_df.iterrows():
        reason = row["exit_reason"]
        r_mean = row["r_mean"]
        win_rate = row["win_rate_pct"]

        if reason == "stop_loss":
            if r_mean < -1.3:
                hints.append(
                    f"stop_loss Ø {r_mean:+.2f}R → Stops zu weit "
                    "oder Slippage-Problem"
                )
            else:
                hints.append(
                    f"stop_loss Ø {r_mean:+.2f}R → Stops korrekt "
                    "platziert (sollte ≈ -1.0R sein)"
                )
        elif reason == "take_profit":
            if r_mean < 1.5:
                hints.append(
                    f"take_profit Ø {r_mean:+.2f}R → Targets zu "
                    "konservativ (MFE prüfen)"
                )
            else:
                hints.append(
                    f"take_profit Ø {r_mean:+.2f}R → Targets gut gesetzt"
                )
        elif reason == "trailing_stop":
            if win_rate < 50:
                hints.append(
                    f"trailing_stop WinRate {win_rate:.0f}% → "
                    "Trailing zu eng"
                )
        elif reason == "eod":
            count_pct = row["count_pct"]
            if count_pct > 20:
                hints.append(
                    f"eod {count_pct:.0f}% aller Trades → "
                    "Strategie hält zu wenig über Nacht"
                )
    return hints


def format_exit_reason_stats(
    stats_df: pd.DataFrame,
    total_trades: int = 0,
) -> str:
    """ASCII-Tabelle der Exit-Grund-Statistik."""
    if stats_df.empty:
        return ""
    if total_trades == 0:
        total_trades = int(stats_df["count"].sum())

    w = 70
    lines = [
        "=" * w,
        f"  EXIT-GRUND ANALYSE  ({total_trades} Trades gesamt)",
        "=" * w,
        "  Grund              Count    %  WinRate    PnL Total"
        "    R̄    Ø Tage",
        "  " + "─" * (w - 4),
    ]
    for _, row in stats_df.iterrows():
        reason = str(row["exit_reason"])[:18].ljust(18)
        cnt = int(row["count"])
        pct = row["count_pct"]
        wr = row["win_rate_pct"]
        pnl_t = row["pnl_total"]
        r_m = row["r_mean"]
        hd = row["hold_days_mean"]
        lines.append(
            f"  {reason} {cnt:>5}  {pct:>5.1f}%  {wr:>6.1f}%"
            f"  {pnl_t:>+12,.2f}  {r_m:>+5.2f}  {hd:>5.0f}"
        )
    lines.append("=" * w)

    # Interpretation
    hints = _build_interpretations(stats_df)
    if hints:
        lines.append("  Interpretation:")
        for h in hints:
            lines.append(f"  {h}")
        lines.append("=" * w)

    return "\n".join(lines) + "\n"


# ═══════════════════════════════════════════════════════════════════════════
#  MAE/MFE-Summary + Top/Worst-Trades (für format_tearsheet)
# ═══════════════════════════════════════════════════════════════════════════

def _format_mae_mfe_summary(trades: list[EnrichedTrade]) -> str:
    """MAE/MFE-Abschnitt für den Tearsheet."""
    if not trades:
        return ""
    mae_rs = [t.mae_r for t in trades]
    mfe_rs = [t.mfe_r for t in trades]
    r_mults = [t.r_multiple for t in trades]

    avg_mae = np.mean(mae_rs) if mae_rs else 0.0
    avg_mfe = np.mean(mfe_rs) if mfe_rs else 0.0

    # Effizienz: wie viel % des MFE wurde als PnL realisiert
    total_mfe_usd = sum(
        t.mfe_r * t.initial_risk_usd for t in trades if t.initial_risk_usd > 0
    )
    total_pnl_net = sum(t.pnl_net for t in trades)
    efficiency = (total_pnl_net / total_mfe_usd * 100.0
                  if total_mfe_usd > 0 else 0.0)

    mae_over_1r = sum(1 for m in mae_rs if m > 1.0)
    mae_over_1r_pct = mae_over_1r / len(trades) * 100.0 if trades else 0.0

    lines = [
        "-" * 50,
        "  MAE / MFE ANALYSE",
        "  " + "─" * 42,
        f"  Ø MAE (R):          {avg_mae:+.2f}R",
        f"  Ø MFE (R):          {avg_mfe:+.2f}R",
        f"  Effizienz-Ratio:    {efficiency:>6.1f}%",
        f"  Trades mit MAE > 1R: {mae_over_1r_pct:.1f}%",
    ]
    return "\n".join(lines) + "\n"


def _format_top_worst_trades(
    trades: list[EnrichedTrade], n: int = 5,
) -> str:
    """Top-N und Worst-N Trades nach R-Multiple."""
    if not trades:
        return ""
    sorted_t = sorted(trades, key=lambda t: t.r_multiple, reverse=True)
    top = sorted_t[:n]
    worst = sorted_t[-n:]

    lines = ["-" * 50, "  TOP-5 TRADES (R-Multiple)"]
    for t in top:
        lines.append(
            f"    {t.symbol:<6} {t.r_multiple:>+6.2f}R  "
            f"${t.pnl_net:>+10,.2f}  {t.exit_reason}"
        )
    lines.append("  WORST-5 TRADES (R-Multiple)")
    for t in worst:
        lines.append(
            f"    {t.symbol:<6} {t.r_multiple:>+6.2f}R  "
            f"${t.pnl_net:>+10,.2f}  {t.exit_reason}"
        )
    return "\n".join(lines) + "\n"


# ═══════════════════════════════════════════════════════════════════════════
#  Trade-Export (CSV + Excel)
# ═══════════════════════════════════════════════════════════════════════════

# Spaltenreihenfolge für CSV/Excel Sheet 1
_EXPORT_COLUMNS = [
    "trade_id", "strategy", "symbol",
    "entry_date", "entry_price", "shares", "entry_signal", "entry_reason",
    "stop_at_entry", "initial_risk_r", "initial_risk_usd",
    "atr_at_entry", "vix_at_entry",
    "exit_date", "exit_price", "exit_reason",
    "hold_days", "hold_trading_days",
    "pnl_gross", "pnl_net", "pnl_pct", "r_multiple",
    "commission", "slippage",
    "strength", "ml_confidence",
    "mae_pct", "mfe_pct", "mae_r", "mfe_r",
    "benchmark_return_pct", "alpha_pct",
    "is_partial", "parent_trade_id",
]


def _trades_to_dataframe(trades: list[EnrichedTrade]) -> pd.DataFrame:
    """Konvertiert EnrichedTrade-Liste in DataFrame mit fester Spaltenfolge."""
    if not trades:
        return pd.DataFrame(columns=_EXPORT_COLUMNS)

    rows = []
    for t in trades:
        row = {}
        for col in _EXPORT_COLUMNS:
            val = getattr(t, col, None)
            # Excel unterstützt keine tz-aware datetimes
            if isinstance(val, datetime) and val.tzinfo is not None:
                val = val.replace(tzinfo=None)
            row[col] = val
        rows.append(row)
    return pd.DataFrame(rows, columns=_EXPORT_COLUMNS)


def export_trades(
    trades: list[EnrichedTrade],
    output_dir: Path,
    fmt: str = "both",
    filename_base: str = "trades",
) -> list[Path]:
    """Exportiert EnrichedTrade-Liste als CSV und/oder Excel.

    Args:
        trades: Liste der abgeschlossenen Trades.
        output_dir: Zielverzeichnis.
        fmt: "csv" | "excel" | "both" | "none".
        filename_base: Dateinamen-Prefix.

    Returns:
        Liste der erstellten Dateipfade.
    """
    if fmt == "none" or not trades:
        return []

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    df = _trades_to_dataframe(trades)
    created: list[Path] = []

    if fmt in ("csv", "both"):
        csv_path = output_dir / f"{filename_base}.csv"
        df.to_csv(csv_path, index=False)
        log.info("export.csv_written", path=str(csv_path), trades=len(df))
        created.append(csv_path)

    if fmt in ("excel", "both"):
        xlsx_path = output_dir / f"{filename_base}.xlsx"
        try:
            _write_excel(df, trades, xlsx_path)
            log.info("export.xlsx_written", path=str(xlsx_path),
                     trades=len(df))
            created.append(xlsx_path)
        except ImportError:
            log.warning("export.openpyxl_missing",
                        hint="pip install openpyxl — Fallback auf CSV")
            if fmt == "excel":
                csv_fb = output_dir / f"{filename_base}.csv"
                df.to_csv(csv_fb, index=False)
                created.append(csv_fb)

    return created


def _strip_tz_from_df(df: pd.DataFrame) -> pd.DataFrame:
    """Entfernt Timezone-Info von allen datetime-Spalten (Excel-Compat)."""
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.tz_localize(None)
        elif df[col].dtype == object:
            # Fallback: einzelne datetime-Objekte in object-Spalten
            df[col] = df[col].apply(
                lambda v: v.replace(tzinfo=None)
                if isinstance(v, datetime) and v.tzinfo is not None
                else v
            )
    return df


def _write_excel(
    df: pd.DataFrame,
    trades: list[EnrichedTrade],
    path: Path,
) -> None:
    """Erzeugt Excel mit 4 Sheets + bedingter Formatierung."""
    import openpyxl
    from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # Sheet 1: Trades
        _strip_tz_from_df(df).to_excel(writer, sheet_name="Trades",
                                        index=False)

        # Sheet 2: Exit Statistik
        stats_df = build_exit_reason_stats(trades)
        if not stats_df.empty:
            _strip_tz_from_df(stats_df).to_excel(
                writer, sheet_name="Exit Statistik", index=False,
            )

        # Sheet 3: MAE/MFE Analyse
        mae_mfe_df = _build_mae_mfe_sheet(trades)
        if not mae_mfe_df.empty:
            _strip_tz_from_df(mae_mfe_df).to_excel(
                writer, sheet_name="MAE-MFE Analyse", index=False,
            )

        # Sheet 4: Zeitverlauf
        time_df = _build_time_series_sheet(trades)
        if not time_df.empty:
            _strip_tz_from_df(time_df).to_excel(
                writer, sheet_name="Zeitverlauf", index=False,
            )

    # Formatierung nachträglich anwenden
    wb = openpyxl.load_workbook(path)
    _apply_trades_formatting(wb["Trades"], df)
    wb.save(path)


def _apply_trades_formatting(ws, df: pd.DataFrame) -> None:
    """Bedingte Formatierung für Sheet 'Trades'."""
    from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    # Erste Zeile einfrieren + AutoFilter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    n_rows = len(df) + 1  # +1 Header
    cols = list(df.columns)

    # pnl_net: grün wenn > 0, rot wenn < 0
    if "pnl_net" in cols:
        col_idx = cols.index("pnl_net") + 1
        col_letter = get_column_letter(col_idx)
        cell_range = f"{col_letter}2:{col_letter}{n_rows}"
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="greaterThan", formula=["0"],
                fill=PatternFill(bgColor="C6EFCE"),
                font=Font(color="006100"),
            ),
        )
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="lessThan", formula=["0"],
                fill=PatternFill(bgColor="FFC7CE"),
                font=Font(color="9C0006"),
            ),
        )

    # r_multiple: Farbskala -3 (rot) → 0 (weiß) → +3 (grün)
    if "r_multiple" in cols:
        col_idx = cols.index("r_multiple") + 1
        col_letter = get_column_letter(col_idx)
        cell_range = f"{col_letter}2:{col_letter}{n_rows}"
        ws.conditional_formatting.add(
            cell_range,
            ColorScaleRule(
                start_type="num", start_value=-3, start_color="FF0000",
                mid_type="num", mid_value=0, mid_color="FFFFFF",
                end_type="num", end_value=3, end_color="00B050",
            ),
        )

    # exit_reason: Kategorie-Farben
    _EXIT_COLORS = {
        "stop_loss": "FFC7CE",
        "take_profit": "C6EFCE",
        "trailing_stop": "BDD7EE",
        "eod": "D9D9D9",
    }
    if "exit_reason" in cols:
        col_idx = cols.index("exit_reason") + 1
        col_letter = get_column_letter(col_idx)
        for reason, color in _EXIT_COLORS.items():
            cell_range = f"{col_letter}2:{col_letter}{n_rows}"
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(
                    operator="equal", formula=[f'"{reason}"'],
                    fill=PatternFill(bgColor=color),
                ),
            )


def _build_mae_mfe_sheet(trades: list[EnrichedTrade]) -> pd.DataFrame:
    """Sheet 3: MAE/MFE-Scatter-Daten + Effizienz-Metriken."""
    if not trades:
        return pd.DataFrame()
    rows = []
    for t in trades:
        mfe_usd = t.mfe_r * t.initial_risk_usd if t.initial_risk_usd > 0 else 0.0
        efficiency = (t.pnl_net / mfe_usd * 100.0) if mfe_usd > 0 else 0.0
        risk_adj_mae = (t.mae_r / abs(t.r_multiple)
                        if t.r_multiple != 0 else 0.0)
        rows.append({
            "trade_id": t.trade_id,
            "symbol": t.symbol,
            "exit_reason": t.exit_reason,
            "mae_r": t.mae_r,
            "mfe_r": t.mfe_r,
            "r_multiple": t.r_multiple,
            "pnl_net": t.pnl_net,
            "efficiency_pct": efficiency,
            "risk_adjusted_mae": risk_adj_mae,
        })
    return pd.DataFrame(rows)


def _build_time_series_sheet(trades: list[EnrichedTrade]) -> pd.DataFrame:
    """Sheet 4: Kumulativer PnL + Rolling Win-Rate + Rolling R-Mean."""
    if not trades:
        return pd.DataFrame()
    sorted_t = sorted(trades, key=lambda t: t.exit_date)
    n = len(sorted_t)
    window = min(20, n)

    cum_pnl = 0.0
    rows = []
    for i, t in enumerate(sorted_t):
        cum_pnl += t.pnl_net
        # Rolling-Fenster (letzte 20 Trades)
        start = max(0, i - window + 1)
        recent = sorted_t[start:i + 1]
        wins = sum(1 for r in recent if r.pnl_net > 0)
        rolling_wr = wins / len(recent) * 100.0
        rolling_r = np.mean([r.r_multiple for r in recent])
        rows.append({
            "trade_nr": i + 1,
            "exit_date": t.exit_date,
            "symbol": t.symbol,
            "pnl_net": t.pnl_net,
            "cumulative_pnl": cum_pnl,
            "rolling_win_rate_pct": rolling_wr,
            "rolling_r_mean": rolling_r,
        })
    return pd.DataFrame(rows)
